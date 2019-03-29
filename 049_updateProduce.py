#! /usr/bin/python3 

# -corrects costs in prduce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# the produce types and their updated prices

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# loop through the rows and update the prices.

for rowNum in range(2, int(sheet.max_row)):   # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')